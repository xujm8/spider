from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from smtplib import SMTP_SSL
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
from time import sleep, localtime, time, strftime
from datetime import datetime, timedelta
from configparser import ConfigParser
import pymssql
import _mssql
import uuid
import decimal
from tkinter import Tk,StringVar,Canvas,Label, messagebox
from openpyxl.workbook import Workbook
import os
#生成日期列表
def create_date_list(datestart = None,dateend = None):
    if datestart is None:
        datestart = '2018-01-01'
    if dateend is None:
        dateend = datetime.now().strftime('%Y-%m-%d')

    # 转为日期格式
    datestart=datetime.strptime(datestart,'%Y-%m-%d')
    dateend=datetime.strptime(dateend,'%Y-%m-%d')
    date_list = []
    date_list.append(datestart.strftime('%Y-%m-%d'))
    while datestart<dateend:
        # 日期叠加一天
        datestart+=timedelta(days=+1)
        # 日期转字符串存入列表
        date_list.append(datestart.strftime('%Y-%m-%d'))
    return date_list


#判断是不是工作日
def isWorkday(date):
    datenow = datetime.strptime(date, '%Y-%m-%d').weekday() + 1
    if datenow == 6 or datenow == 7:
        return False
    return True

#计算当月/下月/当季 买-卖值
def calRes(kind, contractList, priceList, rawdict):
    rawdict[str('%s汇总' % kind)] = sum(priceList)
    d = dict(zip(contractList, priceList))
    contractList.sort()
    if len(contractList) == 1:
        rawdict[str('%s当月' % kind)] = d[contractList[0]]
    elif len(contractList) == 2:
        if contractList[1] - contractList[0] == 1:
            rawdict[str('%s当月' % kind)] = d[contractList[0]]
            rawdict[str('%s下月' % kind)] = d[contractList[1]]
        else:
            rawdict[str('%s当月' % kind)] = d[contractList[0]]
            rawdict[str('%s当季' % kind)] = d[contractList[1]]
    else:
        rawdict[str('%s当月' % kind)] = d[contractList[0]]
        rawdict[str('%s下月' % kind)] = d[contractList[1]]
        rawdict[str('%s当季' % kind)] = d[contractList[2]]
    return rawdict


def allRestoExcel(allRes, allKind, allDate):
    #生成原始dict
    rawdict = {}
    for kind in allKind:
        rawdict['%s汇总' % kind] = 0
        rawdict['%s当月' % kind] = 0
        rawdict['%s下月' % kind] = 0
        rawdict['%s当季' % kind] = 0

    bsdict = {}
    for res in allRes:
        if res[0] not in bsdict.keys():
            bsdict[res[0]] = {}
        if res[1][:2] not in bsdict[res[0]]:
            bsdict[res[0]][res[1][:2]] = {}
        if int(res[-3]) - int(res[-1]) != 0:
            bsdict[res[0]][res[1][:2]][int(res[1][2:])] = int(res[-3]) - int(res[-1])
        else:
            bsdict[res[0]][res[1][:2]][int(res[1][2:])] = 0.1

    excelDict = {}
    for date,bs in bsdict.items():
        rawdictCpy = rawdict.copy()
        for kind in bs:
            rawdictCpy = calRes(kind, list(bs[kind].keys()), list(bs[kind].values()), rawdictCpy)
        excelDict[date] = rawdictCpy

    kindOrderList = []
    for str in ['汇总', '当月', '下月', '当季']:
        for kind in ['IF', 'IC', 'IH']:
            if kind in allKind: 
                kindOrderList.append(kind + str)
    # write in excel
    new = Workbook()
    sheet = new.active
    sheet.title = 'data'
    dateOrderList = list(excelDict.keys())
    dateOrderList.sort()
    for col in range(len(kindOrderList) + 1):
        if col == 0:
            _ = sheet.cell(row = 1, column = col + 1, value = u'日期')
        else:
            _ = sheet.cell(row = 1, column = col + 1, value = u'%s'%kindOrderList[col-1])
    for row in range(len(excelDict.keys())):
        for col in range(len(kindOrderList) + 1):
            if col == 0:
                _ = sheet.cell(row = row + 2, column = col + 1, value = u'%s'%dateOrderList[row])
            else:
                _ = sheet.cell(row = row + 2, column = col + 1, value = u'%s'%excelDict[dateOrderList[row]][kindOrderList[col-1]])

    if len(allDate) == 1:
        filename = '%s.xlsx'%allDate[0]
    else:
        filename = '%s~%s.xlsx'%(allDate[0], allDate[-1])
    newWorkBook = new.save(filename)
    return excelDict, filename





def download_one(kind, date, driver, selector, sleeptime):
    dataList = []
    contractList = []
    selector.select_by_value(kind)
    driver.find_element_by_id("actualDate").clear()
    driver.find_element_by_id('actualDate').send_keys(date)
    button = driver.find_element_by_class_name("btn-query")
    button.click()
    sleep(sleeptime)
    allContract = driver.find_elements_by_css_selector('div.IF_first.clearFloat> a')
    allData = driver.find_elements_by_css_selector('table > tbody > tr.if-table-tr')
    if allContract != []:
        contractList = []
        for contract in allContract:
            contractList.append(contract.text[3:])
        for data in allData:
            dataList.append(data.text.split()[1:])
    return dataList, contractList


def genInfoList(date, dataList, contractList):
    resList = []
    for i in range(len(contractList)):
        resList.append([date, contractList[i]] + dataList[i])
    return resList


def driverStart():
    args = ["hide_console", ]
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    prefs = {"profile.managed_default_content_settings.images": 2}
    chrome_options.add_experimental_option("prefs", prefs)
    #driver = webdriver.Chrome(executable_path = r'chromedriver')
    driver = webdriver.Chrome(options=chrome_options, executable_path=r'chromedriver', service_args = args)
    driver.HideCommandPromptWindow = True
    return driver


#下载多品种多日期
def download(driver, sleeptime = 0.1, allKind = ['IF', 'IC'], allDate = ['2019-02-21'], url = r'http://www.cffex.com.cn/ccpm/'):
    driver.get(url)
    selector = Select(driver.find_element_by_id("selectSec"))
    resList = []
    for date in allDate:
        times = 0
        while True:
            flag = True
            resTmp = []
            for kind in allKind:
                try:
                    dataList, contractList = download_one(kind, date, driver, selector, sleeptime * 5 * (times + 1))
                    if contractList == []:
                        flag = False
                    else:
                        resTmp.extend(genInfoList(date, dataList, contractList))
                except:
                    print(date, kind)
                    sleep(10 * sleeptime)
                    dataList, contractList = download_one(kind, date, driver, selector, 100 * sleeptime)
                    if contractList == []:
                        flag = False
                    else:
                        resTmp.extend(genInfoList(date, dataList, contractList))
            times += 1
            if times == 5 or (times == 1 and isWorkday(date) is False and flag is False): break
            if flag == True:
                resList.extend(resTmp)
                break
    return resList

def readFromConf():
    cf = ConfigParser()
    cf.read("setting.conf")
    #read by type
    dbHost = cf.get("db", "db_host")
    dbPort = cf.get("db", "db_port")
    dbUser = cf.get("db", "db_user")
    dbPass = cf.get("db", "db_pass")
    dbDatabase = cf.get("db", "db_db")
    dbTable = cf.get("db", "db_table")
    beginDate = cf.get("date", "beginDate")
    endDate = cf.get("date", "endDate")
    allKind = cf.get("kind", "allKind")
    sleeptime = float(cf.get("delay", "sleeptime"))
    receivers = cf.get("mailRecv", "receiver")
    allKind = allKind.split(',')
    receivers = receivers.split(',')
    if beginDate == 'today':
        beginDate = strftime('%Y-%m-%d', localtime(time()))
    if endDate == 'today':
        endDate = strftime('%Y-%m-%d', localtime(time()))
    dateList = create_date_list(beginDate, endDate)
    return dbHost, dbPort, dbUser, dbPass, dbDatabase, dbTable, beginDate, endDate, allKind, dateList, sleeptime, receivers


def sqlconnect(server, port, user, password, database, table):
    conn = pymssql.connect(server = '%s:%s'%(server,port),user = user, password = password,database = database)
    cursor = conn.cursor()
    if not cursor:
        print('DATABASE CONNECT FAILED!')
    else:
        #cursor.execute()
        print('DATABASE CONNECT SUCCESSFUL!')
        return conn, cursor


def save_to_sqlserver(table, keys, resList, conn, cursor, allDate, allKind):
    kindstr = ""
    for kind in allKind:
        kindstr += kind[1]
    cursor.execute("if exists (select * from %s where rq between '%s' and '%s' and hydm like ('I[%s]%%')) select '1' else select '0'"%(table, allDate[0], allDate[-1], kindstr))
    if cursor.fetchone()[0] == '1':
        cursor.execute("delete from %s where rq between '%s' and '%s' and hydm like ('I[%s]%%')"%(table, allDate[0], allDate[-1], kindstr))

    insert_sql = 'INSERT INTO {table}({keys})'.format(table=table, keys=keys)
    values = '\' union all select \''.join('\',\''.join(res) for res in resList)
    values = 'select \'' + values + "'"
    # 数据库插入操作
    cursor.execute(insert_sql + values)
    conn.commit()
    print('数据库插入成功!')

    return conn, cursor


def sendEmail(excelstr, filename, receivers):
    # 第三方 SMTP 服务
    mail_host = "smtp.qq.com"  # 设置服务器
    mail_user = "942603714@qq.com"  # 用户名
    mail_pass = "lxqdtnspqemobbci"  # 口令
    sender = '1@qq.com'
    subject = "中金所数据%s已爬取"%filename[:-5]
    body = '<p>数据%s</p>'%excelstr  # 定义邮件正文为html格式
    msg = MIMEMultipart()
    msg['from'] = sender
    msg['to'] = "1@qq.com"
    msg['subject'] = subject
    msg.attach(MIMEText(body, 'html', 'utf-8'))
    att1 = MIMEText(open(filename, 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # 这里的filename可以任意写，写什么名字，邮件中显示什么名字
    att1["Content-Disposition"] = 'attachment; filename=' + filename
    msg.attach(att1)

    smtpObj = SMTP_SSL(host=mail_host)
    smtpObj.connect(host=mail_host, port=465)
    smtpObj.ehlo()
    smtpObj.login(mail_user, mail_pass)
    smtpObj.sendmail(sender, receivers, msg.as_string())
    print("邮件发送成功")
    smtpObj.quit()



class Progress(object):
    """docstring for Progress"""

    def __init__(self):
        self.root = Tk()
        #self.root.geometry('245x30')
        self.root.title('爬虫')
        self.var = StringVar()
        self.var.set("0%")

        # 创建一个背景色为白色的矩形
        self.canvas = Canvas(self.root,width = 120,height = 30,bg = "white")
        # 创建一个矩形外边框（距离左边，距离顶部，矩形宽度，矩形高度），线型宽度，颜色
        self.out_line = self.canvas.create_rectangle(5,5,105,25,width = 1,outline = "black")
        self.fill_line = self.canvas.create_rectangle(5, 5, 5, 25, width = 0, fill = "green")
        Label(self.root,textvariable = self.var).grid(row = 0,column = 0)

        self.canvas.grid(row = 0,column = 1,ipadx = 5)
        self.start()
        self.root.mainloop()


    def change_schedule(self, now_schedule,all_schedule):
        self.canvas.coords(self.fill_line, (5, 5, 6 + (now_schedule/all_schedule)*100, 25))
        self.var.set(str(round(now_schedule/all_schedule*100,2)) + '%')
        if round(now_schedule/all_schedule*100,2) == 100.00:
            self.var.set("完成")
        self.root.update()


    def start(self):
        driver = driverStart()
        fill_line = self.canvas.create_rectangle(2, 2, 0, 27, width = 0, fill = "green")
        starttime = datetime.now()
        try:
        	dbHost, dbPort, dbUser, dbPass, dbDatabase, dbTable, beginDate, endDate, allKind, dateList, sleeptime, receivers = readFromConf()
        	allDate = create_date_list(beginDate, endDate)
        except Exception as ex:
            messagebox.showerror("showerror","读取配置文件失败\nError Message:\n%s"%ex)
        times = len(allDate) // 15 + 1
        self.change_schedule(0, times)
        allRes = []
        flag = True
        try:
        	conn, cursor = sqlconnect(dbHost, dbPort, dbUser, dbPass, dbDatabase, dbTable)
        except Exception as ex:
            messagebox.showerror("showerror","数据库连接失败\nError Message:\n%s"%ex)

        for i in range(times):
            partDate = allDate[i * len(allDate) // times : (i + 1) * len(allDate) // times]
            if partDate == []: continue
            resList = []
            while(resList == [] and (datetime.now()-starttime).seconds / 3600 < 2):
                try:
                    resList = download(driver, sleeptime, allKind, partDate)
                except Exception as ex:
                    messagebox.showerror("showerror","下载失败\nError Message:\n%s"%ex)
                if partDate[-1] != strftime('%Y-%m-%d', localtime(time())):
                    break
        
            if resList != []: 
                try:
                    conn, cursor = save_to_sqlserver(dbTable, 'rq,hydm,cjl,cjlzj,bl,blzj,sl,slzj', resList, conn, cursor, allDate, allKind)
                    allRes.extend(resList)
                except Exception as ex:
                    flag = False
                    print('数据库插入失败!')
                    conn.rollback()
                    messagebox.showerror("showerror","数据库插入失败\nError Message:\n%s"%ex)
                try:
                	self.change_schedule(i + 1,times)
                except Exception as ex:
                    messagebox.showerror("showerror","进度条更新失败\nError Message:\n%s"%ex)


            elif partDate[-1] == strftime('%Y-%m-%d', localtime(time())):
                flag = False
        filename = ""
        if flag:
            try:
                excelDict, filename = allRestoExcel(allRes, allKind, allDate)
                sendEmail(str(excelDict), filename, receivers)
            except Exception as ex:
                print("发送邮件失败")
                messagebox.showerror("showerror","发送邮件失败\nError Message:\n%s"%ex)

        else:
            print("错误：没有数据")
            messagebox.showerror("showerror","错误：没有数据")
        try:
            conn.close()
            driver.quit()
            if(os.path.exists(filename)):
                os.remove(filename)
        except:
        	pass
        self.root.destroy()




if __name__ == "__main__":
    Progress()
